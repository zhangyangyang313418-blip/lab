from __future__ import annotations

import argparse
import json
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "outputs" / "jlr-standards" / "JLR标准名称版本号整理.xlsx"
DEFAULT_HTML = ROOT / "outputs" / "jlr-standards" / "JLR标准名称版本号整理.html"

HEADER_ALIASES = {
    "序号": "index",
    "类别": "category",
    "标准编号": "standard_code",
    "标准名称": "title_zh",
    "标准名称 / 标题": "title_en",
    "标准名称（中文）": "title_zh",
    "中文标准名称": "title_zh",
    "Standard Name (English)": "title_en",
    "标准名称（英文）": "title_en",
    "标题": "title_en",
    "版本号": "version",
    "版本": "version",
    "日期": "date",
    "来源数量": "source_count",
    "来源文件": "source_files",
}

REQUIRED_COLUMNS = {"standard_code", "version"}


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def find_header_row(rows: list[tuple[Any, ...]]) -> tuple[int, dict[str, int]]:
    for row_index, row in enumerate(rows):
        mapping: dict[str, int] = {}
        for column_index, value in enumerate(row):
            label = cell_text(value)
            field = HEADER_ALIASES.get(label)
            if field:
                mapping[field] = column_index
        if REQUIRED_COLUMNS.issubset(mapping) and ("title_zh" in mapping or "title_en" in mapping):
            return row_index, mapping
    raise ValueError("未找到包含“标准编号 / 标准名称 / 版本号”的表头行。")


def load_standards_from_excel(excel_path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)
    sheet = workbook["JLR标准整理"] if "JLR标准整理" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    header_row, mapping = find_header_row(rows)

    standards: list[dict[str, str]] = []
    for row in rows[header_row + 1 :]:
        record = {
            "index": str(len(standards) + 1),
            "category": cell_text(row[mapping["category"]]) if "category" in mapping and mapping["category"] < len(row) else "",
            "standard_code": cell_text(row[mapping["standard_code"]]) if mapping["standard_code"] < len(row) else "",
            "title_zh": cell_text(row[mapping["title_zh"]]) if "title_zh" in mapping and mapping["title_zh"] < len(row) else "",
            "title_en": cell_text(row[mapping["title_en"]]) if "title_en" in mapping and mapping["title_en"] < len(row) else "",
            "version": cell_text(row[mapping["version"]]) if mapping["version"] < len(row) else "",
            "date": cell_text(row[mapping["date"]]) if "date" in mapping and mapping["date"] < len(row) else "",
        }
        if not record["title_zh"]:
            record["title_zh"] = record["title_en"]
        if not record["title_en"]:
            record["title_en"] = record["title_zh"]
        if record["standard_code"] or record["title_zh"] or record["title_en"]:
            standards.append(record)
    return standards


def script_safe_json(value: Any) -> str:
    text = json.dumps(value, ensure_ascii=False, indent=2)
    return text.replace("&", "\\u0026").replace("<", "\\u003C").replace(">", "\\u003E")


def build_html(standards: list[dict[str, str]], source_excel: Path) -> str:
    categories = sorted({row["category"] for row in standards if row.get("category")})
    category_options = "\n".join(
        f'<option value="{escape(category)}">{escape(category)}</option>' for category in categories
    )
    data_json = script_safe_json(standards)
    source_name = escape(source_excel.name)

    return f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>JLR 标准名称与版本号整理</title>
  <style>
    :root {{
      color-scheme: light;
      --ink: #172033;
      --muted: #5b6678;
      --line: #d7dee8;
      --surface: #ffffff;
      --soft: #f4f7fb;
      --blue: #1f5f8f;
      --blue-dark: #153f61;
      --green: #2f7d62;
      --amber: #9a6a16;
      --shadow: 0 10px 28px rgba(23, 32, 51, 0.08);
    }}

    * {{ box-sizing: border-box; }}

    body {{
      margin: 0;
      font-family: "Microsoft YaHei", "PingFang SC", "Segoe UI", Arial, sans-serif;
      color: var(--ink);
      background: #f3f6fa;
    }}

    header {{
      background: #ffffff;
      border-bottom: 1px solid var(--line);
    }}

    .wrap {{
      width: min(1280px, calc(100vw - 32px));
      margin: 0 auto;
    }}

    .topbar {{
      display: flex;
      align-items: flex-end;
      justify-content: space-between;
      gap: 20px;
      padding: 18px 0 16px;
    }}

    h1 {{
      margin: 0 0 5px;
      font-size: 23px;
      line-height: 1.2;
      letter-spacing: 0;
    }}

    .meta {{
      margin: 0;
      color: var(--muted);
      font-size: 13px;
      line-height: 1.5;
    }}

    .summary {{
      display: grid;
      grid-template-columns: repeat(3, minmax(96px, 1fr));
      gap: 10px;
      min-width: 330px;
    }}

    .metric {{
      border: 1px solid var(--line);
      background: var(--surface);
      padding: 10px 12px;
      border-radius: 8px;
    }}

    .metric span {{
      display: block;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.2;
    }}

    .metric strong {{
      display: block;
      margin-top: 4px;
      font-size: 20px;
      line-height: 1.1;
    }}

    main {{
      padding: 18px 0 28px;
    }}

    .toolbar {{
      display: grid;
      grid-template-columns: minmax(300px, 1.8fr) minmax(180px, 0.8fr) minmax(180px, 0.8fr) auto;
      gap: 10px;
      align-items: end;
      margin-bottom: 12px;
    }}

    label {{
      display: grid;
      gap: 6px;
      color: var(--muted);
      font-size: 12px;
      font-weight: 600;
    }}

    input,
    select,
    button {{
      height: 38px;
      border: 1px solid var(--line);
      border-radius: 6px;
      background: var(--surface);
      color: var(--ink);
      font: inherit;
      font-size: 14px;
    }}

    input,
    select {{
      width: 100%;
      padding: 0 11px;
    }}

    button {{
      padding: 0 14px;
      cursor: pointer;
      background: var(--blue);
      border-color: var(--blue);
      color: #fff;
      font-weight: 700;
      white-space: nowrap;
    }}

    button:hover {{
      background: var(--blue-dark);
    }}

    .table-shell {{
      overflow: auto;
      background: var(--surface);
      border: 1px solid var(--line);
      border-radius: 8px;
      box-shadow: var(--shadow);
      max-height: calc(100vh - 205px);
    }}

    table {{
      width: 100%;
      min-width: 1080px;
      border-collapse: collapse;
      table-layout: fixed;
    }}

    thead th {{
      position: sticky;
      top: 0;
      z-index: 1;
      background: #204761;
      color: #fff;
      border-bottom: 1px solid #163750;
      padding: 10px 9px;
      text-align: left;
      font-size: 13px;
      line-height: 1.25;
    }}

    th.sortable {{
      cursor: pointer;
      user-select: none;
    }}

    th.sortable::after {{
      content: "↕";
      float: right;
      opacity: 0.75;
      font-size: 11px;
    }}

    th.sorted-asc::after {{ content: "↑"; }}
    th.sorted-desc::after {{ content: "↓"; }}

    tbody td {{
      border-top: 1px solid #e4eaf1;
      padding: 9px;
      vertical-align: top;
      font-size: 13px;
      line-height: 1.35;
      overflow-wrap: anywhere;
    }}

    tbody tr:nth-child(even) {{
      background: #f8fafc;
    }}

    tbody tr:hover {{
      background: #eaf4f0;
    }}

    .col-index {{ width: 68px; text-align: right; color: var(--muted); }}
    .col-category {{ width: 145px; }}
    .col-code {{ width: 155px; font-weight: 700; color: var(--blue-dark); }}
    .col-title {{ width: 520px; }}
    .col-version {{ width: 120px; color: var(--green); font-weight: 700; }}
    .col-date {{ width: 135px; }}

    thead th.col-index,
    thead th.col-category,
    thead th.col-code,
    thead th.col-title,
    thead th.col-version,
    thead th.col-date {{
      color: #f7fbff;
      font-weight: 700;
    }}

    .title-zh {{
      font-weight: 700;
      line-height: 1.3;
    }}

    .title-en {{
      margin-top: 4px;
      color: var(--muted);
      font-size: 12px;
      line-height: 1.35;
    }}

    .empty {{
      padding: 28px;
      text-align: center;
      color: var(--muted);
      background: var(--surface);
      border: 1px dashed var(--line);
      border-radius: 8px;
    }}

    mark {{
      background: #ffe58a;
      color: #301c00;
      padding: 0 2px;
      border-radius: 2px;
    }}

    @media (max-width: 900px) {{
      .topbar {{
        align-items: stretch;
        flex-direction: column;
      }}

      .summary {{
        min-width: 0;
      }}

      .toolbar {{
        grid-template-columns: 1fr;
      }}

      .table-shell {{
        max-height: none;
      }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="wrap topbar">
      <div>
        <h1>JLR 标准名称与版本号整理</h1>
        <p class="meta">Excel：{source_name}。修改后重新运行同步脚本即可更新本页。</p>
      </div>
      <div class="summary" aria-label="统计">
        <div class="metric"><span>总标准数</span><strong id="totalCount">0</strong></div>
        <div class="metric"><span>当前显示</span><strong id="visibleCount">0</strong></div>
        <div class="metric"><span>类别数</span><strong id="categoryCount">0</strong></div>
      </div>
    </div>
  </header>

  <main class="wrap">
    <section class="toolbar" aria-label="筛选">
      <label>
        关键词搜索
        <input id="searchInput" type="search" placeholder="输入标准编号、标准名称、类别或版本号">
      </label>
      <label>
        类别
        <select id="categoryFilter">
          <option value="">全部类别</option>
          {category_options}
        </select>
      </label>
      <label>
        版本
        <select id="versionFilter">
          <option value="">全部版本</option>
        </select>
      </label>
      <button id="resetButton" type="button">重置</button>
    </section>

    <section class="table-shell" aria-label="JLR 标准表">
      <table>
        <thead>
          <tr>
            <th class="sortable col-index" data-key="index">序号</th>
            <th class="sortable col-category" data-key="category">类别</th>
            <th class="sortable col-code" data-key="standard_code">标准编号</th>
            <th class="sortable col-title" data-key="title_zh">标准名称（中 / 英）</th>
            <th class="sortable col-version" data-key="version">版本号</th>
            <th class="sortable col-date" data-key="date">日期</th>
          </tr>
        </thead>
        <tbody id="tableBody"></tbody>
      </table>
    </section>
    <p id="emptyState" class="empty" hidden>没有匹配的标准。请调整搜索词或筛选条件。</p>
  </main>

  <script>
    const standardsData = {data_json};

    const state = {{
      query: "",
      category: "",
      version: "",
      sortKey: "index",
      sortDirection: "asc",
    }};

    const fields = [
      "index",
      "category",
      "standard_code",
      "title_zh",
      "title_en",
      "version",
      "date",
    ];

    const searchInput = document.getElementById("searchInput");
    const categoryFilter = document.getElementById("categoryFilter");
    const versionFilter = document.getElementById("versionFilter");
    const resetButton = document.getElementById("resetButton");
    const tableBody = document.getElementById("tableBody");
    const emptyState = document.getElementById("emptyState");
    const totalCount = document.getElementById("totalCount");
    const visibleCount = document.getElementById("visibleCount");
    const categoryCount = document.getElementById("categoryCount");
    const headers = Array.from(document.querySelectorAll("th.sortable"));

    function normalize(value) {{
      return String(value ?? "").toLowerCase();
    }}

    function escapeHtml(value) {{
      return String(value ?? "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;")
        .replaceAll("'", "&#039;");
    }}

    function highlight(value) {{
      const text = escapeHtml(value);
      const query = state.query.trim();
      if (!query) return text;
      const escapedQuery = query.replace(/[.*+?^${{}}()|[\\]\\\\]/g, "\\\\$&");
      return text.replace(new RegExp(`(${{escapedQuery}})`, "ig"), "<mark>$1</mark>");
    }}

    function compareRows(left, right) {{
      const key = state.sortKey;
      const a = left[key] ?? "";
      const b = right[key] ?? "";
      const numeric = key === "index";
      let result;
      if (numeric) {{
        result = (Number(a) || 0) - (Number(b) || 0);
      }} else {{
        result = String(a).localeCompare(String(b), "zh-Hans-CN", {{ numeric: true, sensitivity: "base" }});
      }}
      return state.sortDirection === "asc" ? result : -result;
    }}

    function filteredRows() {{
      const query = normalize(state.query.trim());
      return standardsData
        .filter((row) => !state.category || row.category === state.category)
        .filter((row) => !state.version || row.version === state.version)
        .filter((row) => {{
          if (!query) return true;
          return fields.some((field) => normalize(row[field]).includes(query));
        }})
        .sort(compareRows);
    }}

    function render() {{
      const rows = filteredRows();
      tableBody.innerHTML = rows.map((row, rowIndex) => {{
        const rowNumber = rowIndex + 1;
        return `
        <tr>
          <td class="col-index">${{highlight(rowNumber)}}</td>
          <td class="col-category">${{highlight(row.category)}}</td>
          <td class="col-code">${{highlight(row.standard_code)}}</td>
          <td class="col-title">
            <div class="title-zh">${{highlight(row.title_zh)}}</div>
            <div class="title-en">${{highlight(row.title_en)}}</div>
          </td>
          <td class="col-version">${{highlight(row.version)}}</td>
          <td class="col-date">${{highlight(row.date)}}</td>
        </tr>
      `;
      }}).join("");

      visibleCount.textContent = String(rows.length);
      emptyState.hidden = rows.length !== 0;
      headers.forEach((header) => {{
        header.classList.toggle("sorted-asc", header.dataset.key === state.sortKey && state.sortDirection === "asc");
        header.classList.toggle("sorted-desc", header.dataset.key === state.sortKey && state.sortDirection === "desc");
      }});
    }}

    function populateVersionFilter() {{
      const versions = Array.from(new Set(standardsData.map((row) => row.version).filter(Boolean)))
        .sort((a, b) => a.localeCompare(b, "zh-Hans-CN", {{ numeric: true, sensitivity: "base" }}));
      versionFilter.insertAdjacentHTML(
        "beforeend",
        versions.map((version) => `<option value="${{escapeHtml(version)}}">${{escapeHtml(version)}}</option>`).join("")
      );
    }}

    searchInput.addEventListener("input", (event) => {{
      state.query = event.target.value;
      render();
    }});

    categoryFilter.addEventListener("change", (event) => {{
      state.category = event.target.value;
      render();
    }});

    versionFilter.addEventListener("change", (event) => {{
      state.version = event.target.value;
      render();
    }});

    resetButton.addEventListener("click", () => {{
      state.query = "";
      state.category = "";
      state.version = "";
      searchInput.value = "";
      categoryFilter.value = "";
      versionFilter.value = "";
      render();
      searchInput.focus();
    }});

    headers.forEach((header) => {{
      header.addEventListener("click", () => {{
        const key = header.dataset.key;
        if (state.sortKey === key) {{
          state.sortDirection = state.sortDirection === "asc" ? "desc" : "asc";
        }} else {{
          state.sortKey = key;
          state.sortDirection = "asc";
        }}
        render();
      }});
    }});

    totalCount.textContent = String(standardsData.length);
    categoryCount.textContent = String(new Set(standardsData.map((row) => row.category).filter(Boolean)).size);
    populateVersionFilter();
    render();
  </script>
</body>
</html>
"""


def build_html_from_excel(excel_path: Path, html_path: Path) -> Path:
    standards = load_standards_from_excel(excel_path)
    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(build_html(standards, excel_path), encoding="utf-8")
    return html_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build searchable JLR standards HTML from the summary Excel file.")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Path to JLR标准名称版本号整理.xlsx")
    parser.add_argument("--html", type=Path, default=DEFAULT_HTML, help="Output HTML path")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    html_path = build_html_from_excel(args.excel, args.html)
    print(html_path)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
